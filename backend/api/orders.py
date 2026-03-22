from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from pydantic import BaseModel
from typing import Optional
import io

from database import get_db
from models import Order, OrderItem, Product, Supplier
from services.suggestions import calculate_suggestions

router = APIRouter(prefix="/orders", tags=["orders"])


class OrderItemIn(BaseModel):
    product_id: str
    quantity: float
    unit_price: float


class OrderCreate(BaseModel):
    supplier_id: str
    week_start: str   # "2024-01-15"
    notes: Optional[str] = None
    items: list[OrderItemIn]


def order_to_dict(order: Order) -> dict:
    return {
        "id": order.id,
        "supplier_id": order.supplier_id,
        "week_start": order.week_start,
        "status": order.status,
        "notes": order.notes,
        "total_cost": order.total_cost,
        "created_at": order.created_at.isoformat() if order.created_at else None,
        "items": [
            {
                "id": item.id,
                "product_id": item.product_id,
                "product_name": item.product.name if item.product else None,
                "product_code": item.product.code if item.product else None,
                "product_unit": item.product.unit if item.product else None,
                "quantity": item.quantity,
                "unit_price": item.unit_price,
                "total_price": item.total_price,
            }
            for item in order.items
        ],
    }


@router.get("/export")
def export_orders_csv(supplier_id: str | None = None, product_name: str | None = None, db: Session = Depends(get_db)):
    """Export orders as Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(Order).options(joinedload(Order.items).joinedload(OrderItem.product))
    if supplier_id:
        q = q.filter(Order.supplier_id == supplier_id)
    orders = q.order_by(Order.week_start.desc()).all()

    if product_name:
        orders = [o for o in orders if any(
            product_name.lower() in (item.product.name or '').lower()
            for item in o.items if item.product
        )]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "הזמנות"

    # Right-to-left sheet
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="7AAE96")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["ספק", "שבוע", "מוצר", "קוד", "יחידה", "כמות", "מחיר יחידה", 'סה"כ', "סטטוס", "הערות"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for order in orders:
        supplier_name = order.supplier.name if order.supplier else ""
        for item in order.items:
            ws.cell(row=row, column=1, value=supplier_name)
            ws.cell(row=row, column=2, value=order.week_start)
            ws.cell(row=row, column=3, value=item.product.name if item.product else "")
            ws.cell(row=row, column=4, value=item.product.code if item.product else "")
            ws.cell(row=row, column=5, value=item.product.unit if item.product else "")
            ws.cell(row=row, column=6, value=item.quantity)
            ws.cell(row=row, column=7, value=item.unit_price)
            ws.cell(row=row, column=8, value=item.total_price)
            ws.cell(row=row, column=9, value="מאושר" if order.status == "confirmed" else "טיוטה")
            ws.cell(row=row, column=10, value=order.notes or "")
            row += 1

    # Auto column width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=orders.xlsx"}
    )


@router.get("/")
def list_orders(supplier_id: str | None = None, product_name: str | None = None, db: Session = Depends(get_db)):
    q = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.product)
    )
    if supplier_id:
        q = q.filter(Order.supplier_id == supplier_id)
    orders = q.order_by(Order.week_start.desc()).all()

    if product_name:
        orders = [o for o in orders if any(
            product_name.lower() in (item.product.name or '').lower()
            for item in o.items if item.product
        )]

    return [order_to_dict(o) for o in orders]


@router.post("/")
def create_order(body: OrderCreate, db: Session = Depends(get_db)):
    order = Order(
        supplier_id=body.supplier_id,
        week_start=body.week_start,
        notes=body.notes,
    )
    db.add(order)
    db.flush()

    for item in body.items:
        oi = OrderItem(
            order_id=order.id,
            product_id=item.product_id,
            quantity=item.quantity,
            unit_price=item.unit_price,
        )
        db.add(oi)

    db.commit()
    db.refresh(order)

    # reload with relationships
    order = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.product)
    ).filter(Order.id == order.id).first()

    return order_to_dict(order)


@router.get("/{order_id}")
def get_order(order_id: str, db: Session = Depends(get_db)):
    order = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.product)
    ).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(404, "Order not found")
    return order_to_dict(order)


@router.patch("/{order_id}/confirm")
def confirm_order(order_id: str, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(404, "Order not found")
    order.status = "confirmed"
    db.commit()
    return {"ok": True}


@router.delete("/{order_id}")
def delete_order(order_id: str, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(404, "Order not found")
    db.delete(order)
    db.commit()
    return {"ok": True}


@router.get("/suggestions/{supplier_id}")
def order_suggestions(supplier_id: str, db: Session = Depends(get_db)):
    """Suggested quantities based on order history."""
    orders = db.query(Order).options(
        joinedload(Order.items)
    ).filter(Order.supplier_id == supplier_id).order_by(Order.week_start).all()

    history = [
        {
            "id": o.id,
            "items": [
                {"product_id": item.product_id, "quantity": item.quantity}
                for item in o.items
            ],
        }
        for o in orders
    ]

    return calculate_suggestions(history)
